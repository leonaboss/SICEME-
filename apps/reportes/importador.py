import openpyxl
import logging
from datetime import datetime
from django.db import transaction, models
from django.db.models import Max
from django.utils import timezone
from django.contrib.contenttypes.models import ContentType

from apps.emergencias.models import MorbilidadEmergencia, MorbilidadEspecialista, PacienteNoAsistido
from apps.ecosonogramas.models import MorbilidadEcosonograma
from apps.reportes.models import Movimiento
from apps.especialistas.models import Especialidad, Especialista, EstadisticaEspecialidad

logger = logging.getLogger(__name__)

def parse_fecha(valor, default=None):
    if not valor:
        return default
    if isinstance(valor, datetime):
        return valor.date()
    val_str = str(valor).strip()
    if not val_str or val_str.lower() in ['none', 'nan', 'null', '']:
        return default
    
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return default

def parse_sexo(valor, default='M'):
    if not valor: return default
    v = str(valor).strip().upper()
    if v in ['M', 'MASCULINO', 'MASC', 'HOMBRE']:
        return 'M'
    elif v in ['F', 'FEMENINO', 'FEM', 'MUJER']:
        return 'F'
    return default

def parse_int(valor, default=0):
    if valor is None: return default
    try:
        return int(float(valor))
    except (ValueError, TypeError):
        return default

def get_column_map(headers):
    """Detecta automáticamente los índices de las columnas basándose en los encabezados"""
    cmap = {}
    if not headers: return cmap
    
    normalized = [str(h).strip().lower() if h else "" for h in headers]
    
    # Mapeo de términos clave a campos lógicos
    keys = {
        'cedula': ['cedula', 'cédula', 'n° de cédula', 'nro cedula', 'id'],
        'nombre': ['nombre', 'apellido', 'paciente', 'nombre y apellido', 'nombre completo'],
        'edad': ['edad', 'años', 'years'],
        'sexo': ['sexo', 'genero', 'género'],
        'dependencia': ['dependencia', 'procedencia', 'lugar'],
        'telefono': ['telefono', 'teléfono', 'celular', 'contacto'],
        'codigo': ['codigo', 'código', 'cod'],
        'medico': ['medico', 'médico', 'especialista', 'doctor', 'dr'],
        'fecha': ['fecha', 'dia', 'date', 'fecha_diagnostico', 'fecha cita'],
        'motivo': ['motivo', 'consulta', 'motivo consulta'],
        'diagnostico': ['diagnostico', 'diagnóstico', 'resultado'],
        'proxima_cita': ['proxima', 'próxima', 'cita', 'proxima cita'],
        'especialidad': ['especialidad', 'servicio'],
        'tipo_eco': ['tipo eco', 'ecosonograma', 'tipo de eco'],
        'planes': ['planes', 'observaciones', 'plan']
    }
    
    for field, terms in keys.items():
        for i, h in enumerate(normalized):
            if any(term == h or term in h for term in terms):
                cmap[field] = i
                break
    return cmap

def sync_bulk_stats_and_movements(objects, tipo_mov, usuario):
    """Sincroniza estadísticas y movimientos para una lista de objetos creados"""
    if not objects: return
    
    ct = ContentType.objects.get_for_model(objects[0].__class__)
    movimientos = []
    stats_to_update = {} # (esp_id, med_id, mes, anio) -> count
    
    # Cache local para evitar miles de get_or_create
    esp_cache = {} # nombre_lower -> obj
    med_cache = {} # (nombre_lower, esp_id) -> obj

    # Preparar datos
    for obj in objects:
        # 1. Preparar Movimiento
        nombre = getattr(obj, 'nombre_apellido', getattr(obj, 'nombre_completo', 'Paciente'))
        
        detalle = ""
        if tipo_mov == 'emergencia':
            detalle = getattr(obj, 'medico', '')
        elif tipo_mov == 'especialista':
            detalle = f"{getattr(obj, 'especialista', '')} - {getattr(obj, 'especialidad', '')}"
        elif tipo_mov == 'no_asistido':
            detalle = f"{getattr(obj, 'medico', '')} - {getattr(obj, 'especialidad', '')}"
        elif tipo_mov == 'ecosonograma':
            detalle = f"{getattr(obj, 'tipo_eco', '')} - {getattr(obj, 'medico', '')}"

        # Determinar la fecha clínica real para el movimiento
        fecha_mov = timezone.now()
        if tipo_mov == 'emergencia':
            fecha_mov = getattr(obj, 'fecha_diagnostico', timezone.now())
        elif tipo_mov == 'ecosonograma':
            fecha_mov = getattr(obj, 'fecha', timezone.now())
        elif tipo_mov == 'no_asistido':
            fecha_mov = getattr(obj, 'fecha_cita', timezone.now())
        elif tipo_mov == 'especialista':
            # Para especialistas usamos su created_at si existe
            fecha_mov = getattr(obj, 'created_at', timezone.now())

        movimientos.append(Movimiento(
            tipo_mov=tipo_mov,
            nombre_display=nombre,
            detalle=detalle,
            usuario=usuario,
            activo=obj.activo,
            created_at=fecha_mov or timezone.now(),
            content_type=ct,
            object_id=obj.pk
        ))
        
        # 2. Preparar Estadísticas (solo para emergencias y especialistas)
        if tipo_mov in ['emergencia', 'especialista'] and obj.activo:
            esp_nombre = ""
            med_nombre = ""
            fecha = None
            
            if tipo_mov == 'emergencia':
                esp_nombre = 'Emergencias'
                med_nombre = getattr(obj, 'medico', '').strip()
                fecha = getattr(obj, 'fecha_diagnostico', None)
            else:
                esp_nombre = getattr(obj, 'especialidad', '').strip()
                med_nombre = getattr(obj, 'especialista', '').strip()
                fecha = getattr(obj, 'created_at', None)
            
            if not fecha: fecha = timezone.now()
            mes, anio = fecha.month, fecha.year
            
            # Usar Cache para Especialidad
            en_lower = esp_nombre.lower()
            if en_lower not in esp_cache:
                esp, _ = Especialidad.objects.get_or_create(
                    nombre__iexact=esp_nombre,
                    defaults={'nombre': esp_nombre.title() if en_lower != 'emergencias' else 'Emergencias'}
                )
                esp_cache[en_lower] = esp
            else:
                esp = esp_cache[en_lower]

            # Usar Cache para Especialista
            mn_lower = med_nombre.lower()
            if (mn_lower, esp.id) not in med_cache:
                med, _ = Especialista.objects.get_or_create(
                    nombre_completo__iexact=med_nombre,
                    defaults={'nombre_completo': med_nombre.title(), 'especialidad': esp}
                )
                med_cache[(mn_lower, esp.id)] = med
            else:
                med = med_cache[(mn_lower, esp.id)]
            
            key = (esp.id, med.id, mes, anio)
            stats_to_update[key] = stats_to_update.get(key, 0) + 1

    # Bulk create movimientos
    Movimiento.objects.bulk_create(movimientos, batch_size=500)
    
    # Update Estadisticas
    for (esp_id, med_id, mes, anio), count in stats_to_update.items():
        stat, created = EstadisticaEspecialidad.objects.get_or_create(
            especialidad_id=esp_id,
            medico_id=med_id,
            mes=mes,
            anio=anio,
            defaults={'total_pacientes': 0}
        )
        stat.total_pacientes += count
        stat.save()

def procesar_importacion_excel(archivo_xlsx, tipo, usuario):
    """
    Lee un archivo excel y crea registros de forma masiva (bulk).
    """
    wb = openpyxl.load_workbook(archivo_xlsx, data_only=True)
    creados_total = 0
    errores_total = []

    hojas_a_procesar = wb.sheetnames if tipo == 'consolidado' else [wb.active.title]

    for nombre_hoja in hojas_a_procesar:
        ws = wb[nombre_hoja]
        filas = list(ws.iter_rows(values_only=True))
        if len(filas) <= 1: continue 
        
        headers = filas[0]
        cmap = get_column_map(headers)
        datos = filas[1:]
        
        tipo_actual = tipo
        if tipo == 'consolidado':
            n = nombre_hoja.lower()
            if 'emergencia' in n: tipo_actual = 'emergencias'
            elif 'especialista' in n: tipo_actual = 'morbilidad_especialista'
            elif 'eco' in n: tipo_actual = 'ecosonogramas'
            elif 'no asistido' in n or 'no_asistido' in n: tipo_actual = 'no_asistidos'
            else: continue

        to_create = []
        try:
            with transaction.atomic():
                for i, row in enumerate(datos, start=2):
                    if not any(row): continue
                    
                    try:
                        # Helper para obtener valor por cmap o index fallback
                        def gv(field, fallback_idx):
                            idx = cmap.get(field, fallback_idx)
                            if idx < len(row): return row[idx]
                            return None

                        # --- LÓGICA DE DISTRIBUCIÓN INTELIGENTE POR FECHA ---
                        hoy = timezone.now()
                        f_raw = None
                        if tipo_actual == 'emergencias': f_raw = gv('fecha', 9)
                        elif tipo_actual == 'ecosonogramas': f_raw = gv('fecha', 10)
                        elif tipo_actual == 'no_asistidos': f_raw = gv('fecha', 6)
                        
                        f_parsed = parse_fecha(f_raw, hoy.date()) if f_raw else hoy.date()
                        es_mes_actual = (f_parsed.year == hoy.year and f_parsed.month == hoy.month)

                        if tipo_actual == 'emergencias':
                            obj = MorbilidadEmergencia(
                                usuario=usuario,
                                activo=es_mes_actual,
                                cedula=str(gv('cedula', 1)).strip() if gv('cedula', 1) else "",
                                nombre_apellido=str(gv('nombre', 2)).strip() if gv('nombre', 2) else "Sin nombre",
                                edad=parse_int(gv('edad', 3)),
                                sexo=parse_sexo(gv('sexo', 4)),
                                dependencia=str(gv('dependencia', 5)).strip() if gv('dependencia', 5) else "",
                                telefono=str(gv('telefono', 6)).strip() if gv('telefono', 6) else "",
                                codigo=str(gv('codigo', 7)).strip() if gv('codigo', 7) else "",
                                medico=str(gv('medico', 8)).strip() if gv('medico', 8) else "No asignado",
                                fecha_diagnostico=parse_fecha(gv('fecha', 9), timezone.now().date())
                            )
                            to_create.append(obj)
                            
                        elif tipo_actual == 'morbilidad_especialista':
                            obj = MorbilidadEspecialista(
                                usuario=usuario,
                                activo=es_mes_actual,
                                nombre_apellido=str(gv('nombre', 1)).strip() if gv('nombre', 1) else "Sin nombre",
                                edad=parse_int(gv('edad', 2)),
                                sexo=parse_sexo(gv('sexo', 3)),
                                motivo_consulta=str(gv('motivo', 4)).strip() if gv('motivo', 4) else "",
                                diagnostico=str(gv('diagnostico', 5)).strip() if gv('diagnostico', 5) else "",
                                proxima_cita=parse_fecha(gv('proxima_cita', 6)),
                                especialista=str(gv('medico', 7)).strip() if gv('medico', 7) else "No asignado",
                                especialidad=str(gv('especialidad', 8)).strip() if gv('especialidad', 8) else "General"
                            )
                            to_create.append(obj)
                            
                        elif tipo_actual == 'no_asistidos':
                            obj = PacienteNoAsistido(
                                usuario=usuario,
                                activo=es_mes_actual,
                                nombre_completo=str(gv('nombre', 1)).strip() if gv('nombre', 1) else "Sin nombre",
                                edad=parse_int(gv('edad', 2)),
                                sexo=parse_sexo(gv('sexo', 3)),
                                medico=str(gv('medico', 4)).strip() if gv('medico', 4) else "No asignado",
                                especialidad=str(gv('especialidad', 5)).strip() if gv('especialidad', 5) else "General",
                                fecha_cita=parse_fecha(gv('fecha', 6), hoy.date())
                            )
                            to_create.append(obj)
                            
                        elif tipo_actual == 'ecosonogramas':
                            obj = MorbilidadEcosonograma(
                                usuario=usuario,
                                activo=es_mes_actual,
                                nombre_apellido=str(gv('nombre', 1)).strip() if gv('nombre', 1) else "Sin nombre",
                                edad=parse_int(gv('edad', 2)),
                                sexo=parse_sexo(gv('sexo', 3)),
                                cedula=str(gv('cedula', 4)).strip() if gv('cedula', 4) else "",
                                procedencia=str(gv('dependencia', 5)).strip() if gv('dependencia', 5) else "",
                                tipo_eco=str(gv('tipo_eco', 6)).strip() if gv('tipo_eco', 6) else "No especificado",
                                numero_cedula=str(gv('cedula', 7)).strip() if gv('cedula', 7) else "",
                                diagnostico=str(gv('diagnostico', 8)).strip() if gv('diagnostico', 8) else "",
                                medico=str(gv('medico', 9)).strip() if gv('medico', 9) else "No asignado",
                                fecha=parse_fecha(gv('fecha', 10), hoy.date()),
                                planes=str(gv('planes', 11)).strip() if gv('planes', 11) else ""
                            )
                            to_create.append(obj)

                        if len(to_create) >= 1000:
                            model_class = type(to_create[0])
                            # Obtener el ID máximo actual para identificar los nuevos registros
                            max_id = model_class.objects.aggregate(Max('id'))['id__max'] or 0
                            
                            # Carga masiva
                            model_class.objects.bulk_create(to_create)
                            
                            # Recuperar los objetos creados con sus IDs (necesario para Movimientos)
                            created_objs = list(model_class.objects.filter(id__gt=max_id, usuario=usuario).order_by('id'))
                            
                            # Sincronizar estadísticas y movimientos
                            tm = tipo_actual
                            if tm == 'morbilidad_especialista': tm = 'especialista'
                            elif tm == 'emergencias': tm = 'emergencia'
                            elif tm == 'no_asistidos': tm = 'no_asistido'
                            elif tm == 'ecosonogramas': tm = 'ecosonograma'
                            
                            sync_bulk_stats_and_movements(created_objs, tm, usuario)
                            creados_total += len(created_objs)
                            to_create = []

                    except Exception as e:
                        errores_total.append(f"Hoja {nombre_hoja}, Fila {i}: {str(e)}")

                if to_create:
                    model_class = type(to_create[0])
                    max_id = model_class.objects.aggregate(Max('id'))['id__max'] or 0
                    
                    model_class.objects.bulk_create(to_create)
                    created_objs = list(model_class.objects.filter(id__gt=max_id, usuario=usuario).order_by('id'))

                    # Determinar el tipo_mov para sync
                    tm = tipo_actual
                    if tm == 'morbilidad_especialista': tm = 'especialista'
                    elif tm == 'emergencias': tm = 'emergencia'
                    elif tm == 'no_asistidos': tm = 'no_asistido'
                    elif tm == 'ecosonogramas': tm = 'ecosonograma'
                    
                    sync_bulk_stats_and_movements(created_objs, tm, usuario)
                    creados_total += len(created_objs)

        except Exception as e:
            errores_total.append(f"Error crítico en hoja {nombre_hoja}: {str(e)}")

    return creados_total, errores_total
