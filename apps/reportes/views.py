"""
SICEME - Vistas de Dashboard y Reportes
Dashboard principal, estadísticas, exportación Excel, API JSON
"""
import json
from datetime import date
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, Value
from django.db.models.functions import TruncMonth, ExtractMonth, ExtractYear
from django.db import models
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
import traceback

logger = logging.getLogger(__name__)

from apps.usuarios.decorators import rol_requerido
from apps.usuarios.models import Usuario, BitacoraAuditoria
from apps.especialistas.models import Especialidad, Especialista, EstadisticaEspecialidad
from apps.emergencias.models import MorbilidadEmergencia, MorbilidadEspecialista, PacienteNoAsistido
from apps.ecosonogramas.models import MorbilidadEcosonograma
from .models import Movimiento


# ═════════════════════════════════════════════
# DASHBOARD PRINCIPAL
# ═════════════════════════════════════════════
@login_required
def dashboard_view(request):
    """Dashboard principal con estadísticas generales"""
    try:
        hoy = timezone.localtime(timezone.now())
        mes_actual = hoy.month
        anio_actual = hoy.year

        es_admin = request.user.rol == 'ADMIN'

        # ── Filtro base según rol ──
        # ADMIN: ve todo (Global)
        # Otros: solo sus propios registros
        filtro_usuario = {} if es_admin else {'usuario': request.user}

        # ── Contadores ──
        total_emergencias = MorbilidadEmergencia.objects.filter(activo=True, **filtro_usuario).count()
        total_especialistas_morb = MorbilidadEspecialista.objects.filter(activo=True, **filtro_usuario).count()
        total_no_asistidos = PacienteNoAsistido.objects.filter(activo=True, **filtro_usuario).count()
        total_ecosonogramas = MorbilidadEcosonograma.objects.filter(activo=True, **filtro_usuario).count()

        emergencias_mes = MorbilidadEmergencia.objects.filter(
            activo=True,
            fecha_diagnostico__month=mes_actual,
            fecha_diagnostico__year=anio_actual,
            **filtro_usuario
        ).count()

        # Total Movimientos
        total_movimientos = (
            total_emergencias +
            total_especialistas_morb +
            total_no_asistidos +
            total_ecosonogramas
        )

        total_pacientes = total_especialistas_morb + total_emergencias

        # ── Últimos registros (filtrados por usuario) ──
        recientes_emergencias = list(
            MorbilidadEmergencia.objects.filter(**filtro_usuario).order_by('-created_at')[:5]
        )
        recientes_especialistas = list(
            MorbilidadEspecialista.objects.filter(**filtro_usuario).order_by('-created_at')[:5]
        )

        for r in recientes_emergencias:
            r.tipo_morb = 'EMERGENCIA'
        for r in recientes_especialistas:
            r.tipo_morb = 'ESPECIALISTA'

        ultimos_registros = sorted(
            recientes_emergencias + recientes_especialistas,
            key=lambda x: x.created_at or hoy,
            reverse=True
        )[:10]

        # ── Top especialidades (ADMIN ve todo; otros solo sus estadísticas) ──
        if es_admin:
            top_especialidades = list(EstadisticaEspecialidad.objects.filter(
                anio=anio_actual
            ).values('especialidad__nombre').annotate(
                total=Sum('total_pacientes')
            ).order_by('-total')[:5])
        else:
            # Para no-admin calculamos directamente desde sus registros
            from django.db.models import Count
            esp_emergencias = list(
                MorbilidadEmergencia.objects.filter(activo=True, **filtro_usuario)
                .values(especialidad__nombre=models.Value('Emergencias', output_field=models.CharField()))
                .annotate(total=Count('id'))
            )
            esp_especialistas = list(
                MorbilidadEspecialista.objects.filter(activo=True, **filtro_usuario)
                .values('especialidad')
                .annotate(total=Count('id'))
            )
            # Normalizar formato
            top_especialidades = []
            for e in esp_especialistas:
                top_especialidades.append({'especialidad__nombre': e['especialidad'], 'total': e['total']})
            for e in esp_emergencias:
                if e['total'] > 0:
                    top_especialidades.append({'especialidad__nombre': 'Emergencias', 'total': e['total']})
            top_especialidades = sorted(top_especialidades, key=lambda x: x['total'], reverse=True)[:5]

        # ── Especialidades en Vigilancia Centinela (Lógica Avanzada para Admin) ──
        especialidades_criticas_raw = Especialidad.objects.filter(es_vigilancia_centinela=True, activo=True)
        especialidades_criticas = []
        
        for esp in especialidades_criticas_raw:
            # Conteo de hoy
            hoy_count = MorbilidadEspecialista.objects.filter(
                activo=True, especialidad=esp.nombre, created_at__date=hoy.date()
            ).count()
            
            # Conteo mes actual
            mes_count = MorbilidadEspecialista.objects.filter(
                activo=True, especialidad=esp.nombre, 
                created_at__month=mes_actual, created_at__year=anio_actual
            ).count()
            
            # Promedio histórico mensual (aprox) para detectar anomalías
            # Si el mes actual supera el promedio de los últimos 3 meses, es una alerta
            # (Simplificado por ahora a una comparación directa)
            estado = 'Normal'
            clase_estado = 'bg-success'
            if mes_count > 20: # Umbral de ejemplo
                estado = 'Atención'
                clase_estado = 'bg-warning text-dark'
            if mes_count > 50:
                estado = 'Crítico'
                clase_estado = 'bg-danger'
                
            especialidades_criticas.append({
                'nombre': esp.nombre,
                'hoy': hoy_count,
                'mes': mes_count,
                'estado': estado,
                'clase_estado': clase_estado
            })
        
        # Contexto
        context = {
            'total_emergencias': total_emergencias,
            'total_especialistas_morb': total_especialistas_morb,
            'total_no_asistidos': total_no_asistidos,
            'total_ecosonogramas': total_ecosonogramas,
            'total_pacientes': total_pacientes,
            'total_movimientos': total_movimientos,
            'emergencias_mes': emergencias_mes,
            'ultimos_registros': ultimos_registros,
            'top_especialidades': top_especialidades,
            'especialidades_criticas': especialidades_criticas,
            'fecha_actual': hoy.strftime('%d/%m/%Y'),
            'es_admin': es_admin,
        }

        return render(request, 'dashboard/dashboard.html', context)
    except Exception as e:
        import traceback
        import logging
        from django.http import HttpResponseServerError
        logger = logging.getLogger('siceme')
        logger.error(f"Error crítico en dashboard_view: {str(e)}")
        logger.error(traceback.format_exc())
        return HttpResponseServerError(
            f"<h1>Error Crítico en el Dashboard</h1>"
            f"<p>Por favor, capture una captura de pantalla de esto para el desarrollador:</p>"
            f"<pre style='background:#f4f4f4;padding:15px;border:1px solid #ccc'>{traceback.format_exc()}</pre>"
        )


# ═════════════════════════════════════════════
# API ENDPOINTS PARA GRÁFICOS (JSON)
# ═════════════════════════════════════════════
@login_required
def api_dashboard_data(request):
    """API que retorna datos del dashboard para Chart.js"""
    anio = int(request.GET.get('anio', timezone.now().year))

    # Filtro según rol: ADMIN ve todo, otros solo sus propios datos
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    # Emergencias por mes (filtradas por usuario)
    emergencias_por_mes = (
        MorbilidadEmergencia.objects
        .filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)
        .annotate(mes=ExtractMonth('fecha_diagnostico'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )
    meses_data = {item['mes']: item['total'] for item in emergencias_por_mes}
    emergencias_mensual = [meses_data.get(m, 0) for m in range(1, 13)]

    # Nombres de meses en español
    MESES_ES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    hoy = timezone.now()
    mes_nombre = MESES_ES[hoy.month - 1]

    # Distribución por especialidad (filtrada por usuario)
    por_especialidad = (
        MorbilidadEspecialista.objects
        .filter(activo=True, created_at__year=anio, **filtro_usuario)
        .values('especialidad')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    total_pacientes_mes = sum(item['total'] for item in por_especialidad)

    # No asistidos por especialidad (filtrado por usuario)
    no_asistidos_esp = (
        PacienteNoAsistido.objects
        .filter(activo=True, fecha_cita__year=anio, **filtro_usuario)
        .values('especialidad')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    # Top médicos (ADMIN ve tabla global; otros ven sus propios registros calculados)
    if es_admin:
        top_medicos = (
            EstadisticaEspecialidad.objects
            .filter(anio=anio)
            .values('medico__nombre_completo', 'especialidad__nombre')
            .annotate(total=Sum('total_pacientes'))
            .order_by('-total')[:10]
        )
        top_medicos_labels = [item['medico__nombre_completo'] for item in top_medicos]
        top_medicos_data = [item['total'] for item in top_medicos]
    else:
        top_medicos_labels = []
        top_medicos_data = []

    # Ecosonogramas por mes (filtrados por usuario)
    ecosonogramas_por_mes = (
        MorbilidadEcosonograma.objects
        .filter(activo=True, fecha__year=anio, **filtro_usuario)
        .annotate(mes=ExtractMonth('fecha'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )
    eco_meses_data = {item['mes']: item['total'] for item in ecosonogramas_por_mes}
    ecosonogramas_mensual = [eco_meses_data.get(m, 0) for m in range(1, 13)]

    # Ecosonogramas por tipo (filtrados por usuario)
    eco_por_tipo = (
        MorbilidadEcosonograma.objects
        .filter(activo=True, fecha__year=anio, **filtro_usuario)
        .values('tipo_eco')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    total_ecosonogramas_anio = sum(item['total'] for item in eco_por_tipo)

    # ── Vigilancia Centinela Data ──
    # Especialidades marcadas como centinela
    especialidades_centinela_ids = Especialidad.objects.filter(es_vigilancia_centinela=True).values_list('nombre', flat=True)
    
    # Pacientes en especialidades centinela
    pacientes_centinela = MorbilidadEspecialista.objects.filter(
        activo=True, 
        created_at__year=anio, 
        especialidad__in=especialidades_centinela_ids,
        **filtro_usuario
    ).count()
    
    pacientes_otros = total_pacientes_mes - pacientes_centinela if total_pacientes_mes > pacientes_centinela else 0

    # Tendencia mensual centinela
    centinela_por_mes = (
        MorbilidadEspecialista.objects
        .filter(activo=True, created_at__year=anio, especialidad__in=especialidades_centinela_ids, **filtro_usuario)
        .annotate(mes=ExtractMonth('created_at'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )
    centinela_meses_data = {item['mes']: item['total'] for item in centinela_por_mes}
    centinela_mensual = [centinela_meses_data.get(m, 0) for m in range(1, 13)]

    data = {
        'mes_nombre': mes_nombre,
        'anio': hoy.year,
        'emergencias_mensual': {
            'labels': ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                       'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'],
            'data': emergencias_mensual,
        },
        'por_especialidad': {
            'labels': [item['especialidad'] for item in por_especialidad],
            'data': [item['total'] for item in por_especialidad],
            'total': total_pacientes_mes,
        },
        'no_asistidos': {
            'labels': [item['especialidad'] for item in no_asistidos_esp],
            'data': [item['total'] for item in no_asistidos_esp],
        },
        'top_medicos': {
            'labels': top_medicos_labels,
            'data': top_medicos_data,
        },
        'ecosonogramas_mensual': {
            'labels': ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                       'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'],
            'data': ecosonogramas_mensual,
        },
        'eco_por_tipo': {
            'labels': [item['tipo_eco'] for item in eco_por_tipo],
            'data': [item['total'] for item in eco_por_tipo],
            'total': total_ecosonogramas_anio,
        },
        'vigilancia_centinela': {
            'labels': ['Vigilancia Centinela', 'Otras Especialidades'],
            'data': [pacientes_centinela, pacientes_otros],
            'mensual': centinela_mensual,
            'nombres_centinela': list(especialidades_centinela_ids)
        }
    }

    return JsonResponse(data)


@login_required
def api_estadisticas_especialidad(request):
    """API para estadísticas tipo stock dinámico"""
    anio = int(request.GET.get('anio', timezone.now().year))
    mes = request.GET.get('mes')
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    if es_admin:
        stats = EstadisticaEspecialidad.objects.select_related(
            'especialidad', 'medico'
        ).filter(anio=anio)

        if mes:
            stats = stats.filter(mes=int(mes))

        data = [{
            'especialidad': s.especialidad.nombre,
            'medico': s.medico.nombre_completo,
            'total_pacientes': s.total_pacientes,
            'mes': s.mes,
            'anio': s.anio,
        } for s in stats]
    else:
        esp_qs = MorbilidadEspecialista.objects.filter(activo=True, created_at__year=anio, **filtro_usuario)
        em_qs = MorbilidadEmergencia.objects.filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)

        if mes:
            esp_qs = esp_qs.filter(created_at__month=int(mes))
            em_qs = em_qs.filter(fecha_diagnostico__month=int(mes))

        esp_qs = esp_qs.annotate(mes_val=ExtractMonth('created_at')).values('especialidad', 'especialista', 'mes_val').annotate(total=Count('id'))
        em_qs = em_qs.annotate(mes_val=ExtractMonth('fecha_diagnostico')).values('medico', 'mes_val').annotate(total=Count('id'))

        data = []
        for s in esp_qs:
            data.append({
                'especialidad': s['especialidad'],
                'medico': s['especialista'],
                'total_pacientes': s['total'],
                'mes': s['mes_val'],
                'anio': anio,
            })
        for s in em_qs:
            data.append({
                'especialidad': 'Emergencias',
                'medico': s['medico'],
                'total_pacientes': s['total'],
                'mes': s['mes_val'],
                'anio': anio,
            })

    return JsonResponse({'estadisticas': data})


# ═════════════════════════════════════════════
# REPORTES DETALLADOS
# ═════════════════════════════════════════════
@login_required
def reporte_especialidades_view(request):
    """Reporte de pacientes por especialidad"""
    anio = int(request.GET.get('anio', timezone.now().year))
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    MESES_ES = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    if es_admin:
        estadisticas_qs = (
            EstadisticaEspecialidad.objects
            .filter(anio=anio)
            .select_related('especialidad', 'medico')
            .order_by('especialidad__nombre', 'medico__nombre_completo', 'mes')
        )
        estadisticas = [
            {
                'especialidad_nombre': s.especialidad.nombre,
                'medico_nombre': s.medico.nombre_completo,
                'mes': s.mes,
                'mes_nombre': MESES_ES.get(s.mes, str(s.mes)),
                'anio': s.anio,
                'total_pacientes': s.total_pacientes,
            }
            for s in estadisticas_qs
        ]
    else:
        esp_qs = (
            MorbilidadEspecialista.objects
            .filter(activo=True, created_at__year=anio, **filtro_usuario)
            .annotate(mes_val=ExtractMonth('created_at'))
            .values('especialidad', 'especialista', 'mes_val')
            .annotate(total_pacientes=Count('id'))
        )
        em_qs = (
            MorbilidadEmergencia.objects
            .filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)
            .annotate(mes_val=ExtractMonth('fecha_diagnostico'))
            .values('medico', 'mes_val')
            .annotate(total_pacientes=Count('id'))
        )
        estadisticas = []
        for s in esp_qs:
            estadisticas.append({
                'especialidad_nombre': s['especialidad'],
                'medico_nombre': s['especialista'],
                'mes': s['mes_val'],
                'mes_nombre': MESES_ES.get(s['mes_val'], str(s['mes_val'])),
                'anio': anio,
                'total_pacientes': s['total_pacientes'],
            })
        for s in em_qs:
            estadisticas.append({
                'especialidad_nombre': 'Emergencias',
                'medico_nombre': s['medico'],
                'mes': s['mes_val'],
                'mes_nombre': MESES_ES.get(s['mes_val'], str(s['mes_val'])),
                'anio': anio,
                'total_pacientes': s['total_pacientes'],
            })
        estadisticas = sorted(estadisticas, key=lambda x: (x['especialidad_nombre'], x['medico_nombre'], x['mes']))

    return render(request, 'reportes/especialidades.html', {
        'estadisticas': estadisticas,
        'anio': anio,
    })


@login_required
def reporte_emergencias_mes_view(request):
    """Reporte de emergencias por mes"""
    anio = int(request.GET.get('anio', timezone.now().year))
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    emergencias = (
        MorbilidadEmergencia.objects
        .filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)
        .annotate(mes=ExtractMonth('fecha_diagnostico'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    MESES_ES = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    emergencias_data = []
    for e in emergencias:
        emergencias_data.append({
            'mes': e['mes'],
            'mes_nombre': MESES_ES.get(e['mes'], str(e['mes'])),
            'especialidad__nombre': 'Emergencias',
            'total': e['total']
        })

    return render(request, 'reportes/emergencias_mes.html', {
        'emergencias': emergencias_data,
        'anio': anio,
    })


@login_required
def reporte_no_asistidos_view(request):
    """Reporte de pacientes no asistidos por especialidad"""
    anio = int(request.GET.get('anio', timezone.now().year))
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    no_asistidos = (
        PacienteNoAsistido.objects
        .filter(activo=True, fecha_cita__year=anio, **filtro_usuario)
        .values('especialidad')
        .annotate(total=Count('id'))
        .order_by('-total')
    )

    return render(request, 'reportes/no_asistidos.html', {
        'no_asistidos': no_asistidos,
        'anio': anio,
    })


@login_required
def reporte_top_medicos_view(request):
    """Reporte de top médicos por cantidad de pacientes"""
    anio = int(request.GET.get('anio', timezone.now().year))
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    if es_admin:
        top_emergencias = (
            EstadisticaEspecialidad.objects
            .filter(anio=anio, especialidad__nombre__iexact='Emergencias')
            .values('medico__nombre_completo', 'especialidad__nombre')
            .annotate(total=Sum('total_pacientes'))
            .order_by('-total')[:10]
        )

        top_especialistas = (
            EstadisticaEspecialidad.objects
            .filter(anio=anio)
            .exclude(especialidad__nombre__iexact='Emergencias')
            .values('medico__nombre_completo', 'especialidad__nombre')
            .annotate(total=Sum('total_pacientes'))
            .order_by('-total')[:20]
        )
    else:
        top_emergencias = (
            MorbilidadEmergencia.objects
            .filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)
            .values(medico__nombre_completo=models.F('medico'), especialidad__nombre=models.Value('Emergencias', output_field=models.CharField()))
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )
        top_especialistas = (
            MorbilidadEspecialista.objects
            .filter(activo=True, created_at__year=anio, **filtro_usuario)
            .values(medico__nombre_completo=models.F('especialista'), especialidad__nombre=models.F('especialidad'))
            .annotate(total=Count('id'))
            .order_by('-total')[:20]
        )

    return render(request, 'reportes/top_medicos.html', {
        'top_emergencias': top_emergencias,
        'top_especialistas': top_especialistas,
        'anio': anio,
    })


# ═════════════════════════════════════════════
# EXPORTACIÓN EXCEL
# ═════════════════════════════════════════════
@login_required
@rol_requerido('ADMIN', 'ESPECIALISTA', 'PUBLICO')
def exportar_excel_view(request):
    """Exportar reportes a Excel"""
    tipo = request.GET.get('tipo', 'emergencias')
    anio = int(request.GET.get('anio', timezone.now().year))
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    wb = openpyxl.Workbook()
    ws = wb.active

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4A148C', end_color='4A148C', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Registro en bitácora
    BitacoraAuditoria.registrar(
        request.user, BitacoraAuditoria.Accion.EXPORTAR,
        f'Exportación Excel: {tipo} ({anio})', request, 'reportes'
    )

    if tipo == 'emergencias':
        ws.title = 'Emergencias'
        # Orden HTML: [Cédula, Nombre y Apellido, Edad, Sexo, Dependencia, Teléfono, Código, Médico, Fecha]
        headers = ['#', 'Cédula', 'Nombre y Apellido', 'Edad', 'Sexo', 'Dependencia', 'Teléfono', 'Código', 'Médico', 'Fecha']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        registros = MorbilidadEmergencia.objects.filter(
            fecha_diagnostico__year=anio, **filtro_usuario
        ).order_by('fecha_diagnostico')

        for i, reg in enumerate(registros, 1):
            ws.cell(row=i+1, column=1, value=i).border = thin_border
            ws.cell(row=i+1, column=2, value=reg.cedula).border = thin_border
            ws.cell(row=i+1, column=3, value=reg.nombre_apellido).border = thin_border
            ws.cell(row=i+1, column=4, value=reg.edad).border = thin_border
            ws.cell(row=i+1, column=5, value=reg.get_sexo_display()).border = thin_border
            ws.cell(row=i+1, column=6, value=reg.dependencia).border = thin_border
            ws.cell(row=i+1, column=7, value=reg.telefono).border = thin_border
            ws.cell(row=i+1, column=8, value=reg.codigo).border = thin_border
            ws.cell(row=i+1, column=9, value=reg.medico).border = thin_border
            ws.cell(row=i+1, column=10, value=str(reg.fecha_diagnostico)).border = thin_border

        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    elif tipo == 'morbilidad_especialista':
        ws.title = 'Especialistas'
        headers = ['#', 'Nombre y Apellido', 'Edad', 'Sexo', 'Motivo Consulta', 'Diagnóstico', 'Próxima Cita', 'Especialista', 'Especialidad']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        registros = MorbilidadEspecialista.objects.filter(
            proxima_cita__year=anio, **filtro_usuario
        ).order_by('proxima_cita')

        for i, reg in enumerate(registros, 1):
            ws.cell(row=i+1, column=1, value=i).border = thin_border
            ws.cell(row=i+1, column=2, value=reg.nombre_apellido).border = thin_border
            ws.cell(row=i+1, column=3, value=reg.edad).border = thin_border
            ws.cell(row=i+1, column=4, value=reg.get_sexo_display()).border = thin_border
            ws.cell(row=i+1, column=5, value=reg.motivo_consulta).border = thin_border
            ws.cell(row=i+1, column=6, value=reg.diagnostico).border = thin_border
            ws.cell(row=i+1, column=7, value=str(reg.proxima_cita) if reg.proxima_cita else "-").border = thin_border
            ws.cell(row=i+1, column=8, value=reg.especialista).border = thin_border
            ws.cell(row=i+1, column=9, value=reg.especialidad).border = thin_border

        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    elif tipo == 'estadisticas':
        ws.title = 'Estadísticas'
        headers = ['#', 'Especialidad', 'Médico', 'Total Pacientes', 'Mes', 'Año']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        if es_admin:
            stats = EstadisticaEspecialidad.objects.filter(
                anio=anio
            ).select_related('especialidad', 'medico').order_by('especialidad__nombre')

            for i, s in enumerate(stats, 1):
                ws.cell(row=i+1, column=1, value=i).border = thin_border
                ws.cell(row=i+1, column=2, value=s.especialidad.nombre).border = thin_border
                ws.cell(row=i+1, column=3, value=s.medico.nombre_completo).border = thin_border
                ws.cell(row=i+1, column=4, value=s.total_pacientes).border = thin_border
                ws.cell(row=i+1, column=5, value=s.mes).border = thin_border
                ws.cell(row=i+1, column=6, value=s.anio).border = thin_border
        else:
            esp_qs = (
                MorbilidadEspecialista.objects
                .filter(activo=True, created_at__year=anio, **filtro_usuario)
                .annotate(mes_val=ExtractMonth('created_at'))
                .values('especialidad', 'especialista', 'mes_val')
                .annotate(total_pacientes=Count('id'))
            )
            em_qs = (
                MorbilidadEmergencia.objects
                .filter(activo=True, fecha_diagnostico__year=anio, **filtro_usuario)
                .annotate(mes_val=ExtractMonth('fecha_diagnostico'))
                .values('medico', 'mes_val')
                .annotate(total_pacientes=Count('id'))
            )
            data_list = []
            for s in esp_qs:
                data_list.append((s['especialidad'], s['especialista'], s['total_pacientes'], s['mes_val'], anio))
            for s in em_qs:
                data_list.append(('Emergencias', s['medico'], s['total_pacientes'], s['mes_val'], anio))
            
            data_list.sort(key=lambda x: (x[0], x[1], x[3]))

            for i, row_data in enumerate(data_list, 1):
                ws.cell(row=i+1, column=1, value=i).border = thin_border
                ws.cell(row=i+1, column=2, value=row_data[0]).border = thin_border
                ws.cell(row=i+1, column=3, value=row_data[1]).border = thin_border
                ws.cell(row=i+1, column=4, value=row_data[2]).border = thin_border
                ws.cell(row=i+1, column=5, value=row_data[3]).border = thin_border
                ws.cell(row=i+1, column=6, value=row_data[4]).border = thin_border

        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    elif tipo == 'no_asistidos':
        ws.title = 'No Asistidos'
        # Orden HTML: [Nombre Completo, Edad, Sexo, Médico, Especialidad, Fecha Cita]
        headers = ['#', 'Nombre Completo', 'Edad', 'Sexo', 'Médico', 'Especialidad', 'Fecha Cita']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        registros = PacienteNoAsistido.objects.filter(
            fecha_cita__year=anio, **filtro_usuario
        ).order_by('fecha_cita')

        for i, reg in enumerate(registros, 1):
            ws.cell(row=i+1, column=1, value=i).border = thin_border
            ws.cell(row=i+1, column=2, value=reg.nombre_completo).border = thin_border
            ws.cell(row=i+1, column=3, value=reg.edad).border = thin_border
            ws.cell(row=i+1, column=4, value=reg.get_sexo_display()).border = thin_border
            ws.cell(row=i+1, column=5, value=reg.medico).border = thin_border
            ws.cell(row=i+1, column=6, value=reg.especialidad).border = thin_border
            ws.cell(row=i+1, column=7, value=str(reg.fecha_cita)).border = thin_border

        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    elif tipo == 'ecosonogramas':
        ws.title = 'Ecosonogramas'
        # Orden HTML: [Nombre y Apellido, Edad, Sexo, Cédula, Procedencia, Tipo Eco, N° de Cédula, Diagnóstico, Médico, Fecha, Planes]
        headers = ['#', 'Nombre y Apellido', 'Edad', 'Sexo', 'Cédula', 'Procedencia', 'Tipo Eco', 'N° de Cédula', 'Diagnóstico', 'Médico', 'Fecha', 'Planes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        registros = MorbilidadEcosonograma.objects.filter(
            fecha__year=anio, **filtro_usuario
        ).order_by('fecha')

        for i, reg in enumerate(registros, 1):
            ws.cell(row=i+1, column=1, value=i).border = thin_border
            ws.cell(row=i+1, column=2, value=reg.nombre_apellido).border = thin_border
            ws.cell(row=i+1, column=3, value=reg.edad).border = thin_border
            ws.cell(row=i+1, column=4, value=reg.get_sexo_display()).border = thin_border
            ws.cell(row=i+1, column=5, value=reg.cedula).border = thin_border
            ws.cell(row=i+1, column=6, value=reg.procedencia).border = thin_border
            ws.cell(row=i+1, column=7, value=reg.tipo_eco).border = thin_border
            ws.cell(row=i+1, column=8, value=reg.numero_cedula).border = thin_border
            ws.cell(row=i+1, column=9, value=reg.diagnostico).border = thin_border
            ws.cell(row=i+1, column=10, value=reg.medico).border = thin_border
            ws.cell(row=i+1, column=11, value=str(reg.fecha)).border = thin_border
            ws.cell(row=i+1, column=12, value=reg.planes).border = thin_border

        for col in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="SICEME_{tipo}_{anio}.xlsx"'
    wb.save(response)
    return response

# ═════════════════════════════════════════════
# IMPORTACIÓN EXCEL
# ═════════════════════════════════════════════
from django.shortcuts import redirect
from .importador import procesar_importacion_excel

@login_required
@rol_requerido('ADMIN', 'ESPECIALISTA')
def importar_excel_view(request):
    """Importar datos masivos desde Excel"""
    if request.method == 'POST':
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'No se proporcionó ningún archivo.')
            return redirect(request.META.get('HTTP_REFERER', '/'))
            
        archivo = request.FILES['archivo_excel']
        tipo = request.POST.get('tipo', 'emergencias')
        
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser de tipo Excel (.xlsx).')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Registrar en bitácora
        BitacoraAuditoria.registrar(
            request.user, BitacoraAuditoria.Accion.CREAR,
            f'Importación masiva Excel: {tipo}', request, 'reportes'
        )

        creados, errores = procesar_importacion_excel(archivo, tipo, request.user)
        
        if creados > 0:
            messages.success(request, f'Se han importado {creados} registros correctamente.')
        
        if errores:
            # Mostramos un resumen de errores en un alert warning
            msg_errores = f"Se omitieron registros con errores ({len(errores)}). Primeros fallos: " + " | ".join(errores[:3])
            messages.warning(request, msg_errores)

        return redirect(request.META.get('HTTP_REFERER', '/'))
        
    return redirect('dashboard')



# ═════════════════════════════════════════════
# MOVIMIENTOS - Panel de Actividad + Archivados
# ═════════════════════════════════════════════
@login_required
def movimientos_view(request):
    """Panel de movimientos: timeline de actividad + registros archivados (Usando tabla central Movimiento)"""
    import datetime as dt
    hoy = timezone.now()
    inicio_mes = dt.date(hoy.year, hoy.month, 1)
    if hoy.month == 12:
        fin_mes = dt.date(hoy.year + 1, 1, 1)
    else:
        fin_mes = dt.date(hoy.year, hoy.month + 1, 1)

    # Filtrar por usuario si no es admin
    filtro_usuario = {} if request.user.rol == 'ADMIN' else {'usuario': request.user}

    # ── Tarjetas de resumen (mes actual) ──
    movimientos_mes = Movimiento.objects.filter(
        activo=True, created_at__gte=inicio_mes, created_at__lt=fin_mes, **filtro_usuario
    )
    
    emergencias_mes = movimientos_mes.filter(tipo_mov='emergencia').count()
    especialistas_mes = movimientos_mes.filter(tipo_mov='especialista').count()
    ecosonogramas_mes = movimientos_mes.filter(tipo_mov='ecosonograma').count()
    no_asistidos_mes = movimientos_mes.filter(tipo_mov='no_asistido').count()
    
    pacientes_atendidos = emergencias_mes + especialistas_mes + ecosonogramas_mes
    total_archivados = Movimiento.objects.filter(activo=False, **filtro_usuario).count()

    # ── Timeline de actividad reciente (últimos 20 movimientos activos) ──
    timeline = Movimiento.objects.filter(activo=True, **filtro_usuario).order_by('-created_at')[:20]

    # ── Registros Archivados Recientemente (Papelera) ──
    q_archivados = request.GET.get('q_archivados', '')
    archivados = Movimiento.objects.filter(activo=False, **filtro_usuario).order_by('-created_at')
    
    if q_archivados:
        archivados = archivados.filter(nombre_display__icontains=q_archivados)
    
    # Limitamos a los últimos 50 para rendimiento
    archivados = archivados[:50]

    MESES_LISTA = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    MESES_DICT = dict(MESES_LISTA)

    context = {
        'pacientes_atendidos': pacientes_atendidos,
        'no_asistidos_mes': no_asistidos_mes,
        'total_archivados': total_archivados,
        'emergencias_mes': emergencias_mes,
        'especialistas_mes': especialistas_mes,
        'ecosonogramas_mes': ecosonogramas_mes,
        'registros_activos_mes': pacientes_atendidos,
        'timeline': timeline,
        'archivados': archivados,
        'q_archivados': q_archivados,
        'mes_nombre': MESES_DICT.get(hoy.month),
        'anio': hoy.year,
    }
    return render(request, 'reportes/movimientos.html', context)

@login_required
@rol_requerido('ADMIN')
def biblioteca_view(request):
    """
    Vista de la Biblioteca Histórica (Solo Administrador).
    Muestra los cierres de mes realizados y permite navegar por ellos.
    """
    from .models import CierreMes
    
    cierres = CierreMes.objects.all()
    
    # Detección de meses para carpetas (Forma Robusta compatible con cualquier MySQL)
    from django.db.models import Count
    movimientos_archivados = Movimiento.objects.filter(activo=False).only('created_at')
    
    biblioteca_estructurada = {}
    MESES_LISTA = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    MESES_DICT = dict(MESES_LISTA)

    for m in movimientos_archivados:
        if m.created_at:
            anio = str(m.created_at.year)
            mes_num = m.created_at.month
            
            if anio not in biblioteca_estructurada:
                biblioteca_estructurada[anio] = {}
            
            if mes_num not in biblioteca_estructurada[anio]:
                biblioteca_estructurada[anio][mes_num] = {
                    'mes_num': mes_num,
                    'mes_nombre': MESES_DICT.get(mes_num, 'Desconocido'),
                    'total': 0
                }
            
            biblioteca_estructurada[anio][mes_num]['total'] += 1

    # Convertir los diccionarios internos a listas ordenadas para el template
    for anio in biblioteca_estructurada:
        meses_lista_anio = list(biblioteca_estructurada[anio].values())
        # Ordenar meses de mayor a menor
        meses_lista_anio.sort(key=lambda x: x['mes_num'], reverse=True)
        biblioteca_estructurada[anio] = meses_lista_anio

    # Filtros para la lista detallada
    mes_filtro = request.GET.get('mes')
    anio_filtro = request.GET.get('anio')
    q = request.GET.get('q', '')
    
    archivados = Movimiento.objects.filter(activo=False).order_by('-created_at')
    
    if mes_filtro:
        archivados = archivados.filter(created_at__month=mes_filtro)
    if anio_filtro:
        archivados = archivados.filter(created_at__year=anio_filtro)
    if q:
        archivados = archivados.filter(nombre_display__icontains=q)

    context = {
        'cierres': cierres,
        'biblioteca': biblioteca_estructurada,
        'archivados': archivados,
        'meses_lista': MESES_LISTA,
        'filtros': {
            'mes': int(mes_filtro) if mes_filtro else None,
            'anio': anio_filtro,
            'q': q,
            'mes_nombre': MESES_DICT.get(int(mes_filtro)) if mes_filtro else None
        }
    }
    return render(request, 'reportes/biblioteca.html', context)

@login_required
@rol_requerido('ADMIN', 'ESPECIALISTA')
def restaurar_masivo_view(request):
    """Restaurar una cantidad específica de registros archivados recientemente."""
    from django.db import transaction
    if request.method == 'POST':
        cantidad_str = request.POST.get('cantidad')
        try:
            cantidad = int(cantidad_str)
            if cantidad <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messages.error(request, 'Por favor, ingrese una cantidad válida de registros a restaurar.')
            return redirect(request.META.get('HTTP_REFERER', 'movimientos'))
            
        # Filtro de usuario si no es admin
        filtro_usuario = {}
        if request.user.rol != 'ADMIN':
            filtro_usuario['usuario'] = request.user
            
        # Obtener los últimos N movimientos archivados
        movs_a_restaurar = Movimiento.objects.filter(activo=False, **filtro_usuario).order_by('-created_at')[:cantidad]
        total = movs_a_restaurar.count()
        
        if total == 0:
            messages.info(request, 'No hay registros archivados para restaurar.')
            return redirect(request.META.get('HTTP_REFERER', 'movimientos'))
            
        with transaction.atomic():
            for mov in movs_a_restaurar:
                # Restaurar el original
                obj = mov.registro_original
                if obj:
                    obj.activo = True
                    obj.save()
                # Restaurar el movimiento
                mov.activo = True
                mov.save()
        
        BitacoraAuditoria.registrar(
            request.user, BitacoraAuditoria.Accion.EDITAR,
            f'Restauración masiva realizada: {total} registros recuperados', request, 'reportes'
        )
        messages.success(request, f'¡Éxito! Se han restaurado los últimos {total} registros correctamente.')
        
    return redirect(request.META.get('HTTP_REFERER', 'movimientos'))

@login_required
@rol_requerido('ADMIN')
def cerrar_mes_view(request):
    """
    Acción para realizar el cierre formal de un mes.
    Archiva todos los registros activos del mes/año seleccionado.
    """
    from .models import CierreMes
    if request.method == 'POST':
        mes = int(request.POST.get('mes'))
        anio = int(request.POST.get('anio'))
        
        # 1. Identificar registros activos de ese mes
        # Buscamos en todos los modelos de morbilidad
        m_emergencia = MorbilidadEmergencia.objects.filter(activo=True, fecha_diagnostico__month=mes, fecha_diagnostico__year=anio)
        m_especialista = MorbilidadEspecialista.objects.filter(activo=True, created_at__month=mes, created_at__year=anio)
        m_no_asistido = PacienteNoAsistido.objects.filter(activo=True, created_at__month=mes, created_at__year=anio)
        m_eco = MorbilidadEcosonograma.objects.filter(activo=True, fecha__month=mes, fecha__year=anio)
        
        total = m_emergencia.count() + m_especialista.count() + m_no_asistido.count() + m_eco.count()
        
        if total == 0:
            messages.warning(request, f'No se encontraron registros activos para el mes {mes}/{anio}.')
            return redirect('biblioteca')
            
        with transaction.atomic():
            # 2. Archivar (Soft Delete)
            m_emergencia.update(activo=False)
            m_especialista.update(activo=False)
            m_no_asistido.update(activo=False)
            m_eco.update(activo=False)
            
            # Sincronizar Movimientos (asegurar que activo=False)
            Movimiento.objects.filter(
                created_at__month=mes, created_at__year=anio
            ).update(activo=False)
            
            # 3. Crear el registro de Cierre
            CierreMes.objects.update_or_create(
                mes=mes, anio=anio,
                defaults={
                    'usuario_cierre': request.user,
                    'total_registros': total
                }
            )
            
            BitacoraAuditoria.registrar(
                request.user, BitacoraAuditoria.Accion.EDITAR,
                f'Cierre Mensual formalizado: {mes}/{anio} ({total} registros)', request, 'reportes'
            )
            messages.success(request, f'¡Cierre de mes {mes}/{anio} completado con éxito! {total} registros movidos a la Biblioteca.')
        
    return redirect('biblioteca')

@login_required
@rol_requerido('ADMIN')
def auto_organizar_biblioteca_view(request):
    """
    Escanea TODOS los registros (activos e inactivos) y los organiza en la biblioteca 
    según su fecha clínica real. Corrige carpetas mal ubicadas.
    """
    from django.db import transaction
    from django.contrib.contenttypes.models import ContentType
    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    with transaction.atomic():
        def reindexar_tipo(model_class, campo_fecha):
            # Procesamos TODO (tanto activos como ya archivados)
            registros = model_class.objects.all()
            count_cambios = 0
            ct = ContentType.objects.get_for_model(model_class)
            tipo_mov = ''
            if model_class == MorbilidadEmergencia: tipo_mov = 'emergencia'
            elif model_class == MorbilidadEcosonograma: tipo_mov = 'ecosonograma'
            elif model_class == MorbilidadEspecialista: tipo_mov = 'especialista'
            elif model_class == PacienteNoAsistido: tipo_mov = 'no_asistido'
            
            for obj in registros:
                fecha_real = getattr(obj, campo_fecha)
                if not fecha_real: continue
                
                # Normalizar fecha_real a datetime si es date para la comparación
                from datetime import date, datetime
                if isinstance(fecha_real, date) and not isinstance(fecha_real, datetime):
                    fecha_comp = datetime.combine(fecha_real, datetime.min.time())
                    # Hacer aware si el sistema usa zonas horarias
                    from django.utils import timezone
                    if timezone.is_aware(inicio_mes):
                        fecha_comp = timezone.make_aware(fecha_comp)
                else:
                    fecha_comp = fecha_real

                # 1. Si es antiguo y estaba activo, lo archivamos
                if obj.activo and fecha_comp < inicio_mes:
                    obj.activo = False
                    obj.save()
                
                # 2. Sincronizar el Movimiento
                # Calculamos el detalle actual según la lógica de señales
                nombre_display = getattr(obj, 'nombre_apellido', getattr(obj, 'nombre_completo', 'Sin nombre'))
                detalle = ''
                if tipo_mov == 'emergencia':
                    detalle = f"Diag: {getattr(obj, 'diagnostico', 'Sin diagnóstico')} | Med: {obj.medico}"
                elif tipo_mov == 'especialista':
                    detalle = f"Diag: {obj.diagnostico} | {obj.especialista} - {obj.especialidad}"
                elif tipo_mov == 'ecosonograma':
                    detalle = f"Diag: {obj.diagnostico} | {obj.tipo_eco} - {obj.medico}"
                elif tipo_mov == 'no_asistido':
                    detalle = f"{obj.medico} - {obj.especialidad}"

                # Buscamos o creamos el movimiento
                mov, created = Movimiento.objects.get_or_create(
                    content_type=ct, object_id=obj.pk,
                    defaults={
                        'tipo_mov': tipo_mov,
                        'nombre_display': nombre_display,
                        'detalle': detalle,
                        'usuario': getattr(obj, 'usuario', None),
                        'activo': obj.activo,
                        'created_at': fecha_real
                    }
                )

                if not created:
                    # Si ya existía, verificamos si hay que actualizar
                    if (mov.created_at.year != fecha_real.year or 
                        mov.created_at.month != fecha_real.month or 
                        mov.activo != obj.activo or
                        mov.detalle != detalle or
                        mov.nombre_display != nombre_display):
                        
                        mov.created_at = fecha_real
                        mov.activo = obj.activo
                        mov.detalle = detalle
                        mov.nombre_display = nombre_display
                        mov.save()
                        count_cambios += 1
                else:
                    count_cambios += 1

            return count_cambios

        c1 = reindexar_tipo(MorbilidadEmergencia, 'fecha_diagnostico')
        c2 = reindexar_tipo(MorbilidadEcosonograma, 'fecha')
        c3 = reindexar_tipo(MorbilidadEspecialista, 'created_at')
        c4 = reindexar_tipo(PacienteNoAsistido, 'fecha_cita')
        
        total = c1 + c2 + c3 + c4
        
        if total > 0:
            BitacoraAuditoria.registrar(
                request.user, BitacoraAuditoria.Accion.EDITAR,
                f'Re-indexación completa de Biblioteca: {total} movimientos corregidos', request, 'reportes'
            )
            messages.success(request, f'¡Sincronización completa! Se han re-organizado {total} registros en sus meses originales.')
        else:
            messages.info(request, 'Toda la biblioteca ya está correctamente organizada por fechas.')
            
    return redirect('biblioteca')


@login_required
@rol_requerido('ADMIN', 'ESPECIALISTA', 'PUBLICO')
def restaurar_registro_view(request):
    """Restaurar un registro archivado a través de su Movimiento central"""
    if request.method == 'POST':
        pk_movimiento = request.POST.get('pk')

        try:
            # Ahora buscamos el movimiento central
            movimiento = Movimiento.objects.get(pk=pk_movimiento, activo=False)

            # Refuerzo de Seguridad: Solo el dueño o el admin pueden restaurar
            if request.user.rol != 'ADMIN' and movimiento.usuario != request.user:
                messages.error(request, 'No tiene permiso para restaurar este registro.')
                return redirect('movimientos')

            registro_original = movimiento.registro_original
            
            if registro_original:
                registro_original.activo = True
                registro_original.save() # Esto disparará la señal post_save y actualizará el Movimiento
                
                BitacoraAuditoria.registrar(
                    request.user, BitacoraAuditoria.Accion.EDITAR,
                    f'Registro restaurado ({movimiento.tipo_mov}): {movimiento.nombre_display}', 
                    request, 'movimientos'
                )
                messages.success(request, 'Registro restaurado exitosamente.')
            else:
                messages.error(request, 'No se pudo localizar el registro original para restaurar.')
                
        except Movimiento.DoesNotExist:
            messages.error(request, 'El registro de movimiento no existe o ya está activo.')
        except Exception as e:
            messages.error(request, f'Error crítico al restaurar: {str(e)}')

    return redirect('movimientos')


@login_required
@rol_requerido('ADMIN', 'ESPECIALISTA')
def limpiar_actividad_global_view(request):
    """Archiva todos los registros activos del sistema (Limpieza Global)."""
    if request.method == 'POST':
        filtro_usuario = {} if request.user.rol == 'ADMIN' else {'usuario': request.user}
        
        # Archivar (setear activo=False) en los 4 modelos
        MorbilidadEmergencia.objects.filter(activo=True, **filtro_usuario).update(activo=False)
        MorbilidadEspecialista.objects.filter(activo=True, **filtro_usuario).update(activo=False)
        PacienteNoAsistido.objects.filter(activo=True, **filtro_usuario).update(activo=False)
        MorbilidadEcosonograma.objects.filter(activo=True, **filtro_usuario).update(activo=False)
        
        # IMPORTANTE: Sincronizar también la tabla de Movimientos (ya que .update() no dispara señales)
        Movimiento.objects.filter(activo=True, **filtro_usuario).update(activo=False)
        
        BitacoraAuditoria.registrar(
            request.user, BitacoraAuditoria.Accion.EDITAR,
            'Limpieza global de historial (Archivados todos los registros activos)', 
            request, 'movimientos'
        )
        messages.success(request, 'El historial de actividad se ha limpiado correctamente.')

    return redirect('movimientos')


@login_required
@rol_requerido('ADMIN')
def limpiar_archivados_view(request):
    """Borra definitivamente todos los registros que ya están archivados (activo=False)."""
    if request.method == 'POST':
        # Borrado físico de la base de datos
        count1 = MorbilidadEmergencia.objects.filter(activo=False).delete()[0]
        count2 = MorbilidadEspecialista.objects.filter(activo=False).delete()[0]
        count3 = PacienteNoAsistido.objects.filter(activo=False).delete()[0]
        count4 = MorbilidadEcosonograma.objects.filter(activo=False).delete()[0]
        
        total = count1 + count2 + count3 + count4
        
        BitacoraAuditoria.registrar(
            request.user, BitacoraAuditoria.Accion.ELIMINAR,
            f'Papelera vaciada: {total} registros eliminados definitivamente', 
            request, 'movimientos'
        )
        messages.success(request, f'Se han eliminado definitivamente {total} registros de la papelera.')

    return redirect('movimientos')


# Las vistas de eliminación definitiva han sido removidas para asegurar la integridad de los datos médicos.
# Se utiliza Soft Delete (archivado) en todo el sistema.





@login_required
def reporte_ecosonogramas_enfermedades_view(request):
    anio_actual = timezone.now().year
    anio = request.GET.get('anio', anio_actual)
    try:
        anio = int(anio)
    except ValueError:
        anio = anio_actual

    registros = MorbilidadEcosonograma.objects.filter(activo=True, fecha__year=anio)
    if request.user.rol != 'ADMIN':
         registros = registros.filter(usuario=request.user)

    enfermedades = (
        registros
        .values('diagnostico')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    return render(request, 'reportes/ecosonogramas_enfermedades.html', {
        'enfermedades': enfermedades,
        'anio': anio
    })


@login_required
@rol_requerido('ADMIN')
def monitor_view(request):
    """Vista dedicada para el monitoreo de actividad en tiempo real"""
    actividad = Movimiento.objects.filter(activo=True).select_related('usuario').order_by('-created_at')[:50]
    return render(request, 'reportes/monitor.html', {
        'actividad': actividad,
        'fecha_actual': timezone.localtime(timezone.now()).strftime('%d/%m/%Y')
    })


@login_required
def reporte_periodo_view(request):
    """Generador de reportes con desglose mensual y gráficos para periodos personalizados"""
    hoy = timezone.now()
    mes_inicio = int(request.GET.get('mes_inicio', 1))
    mes_fin = int(request.GET.get('mes_fin', hoy.month))
    anio = int(request.GET.get('anio', hoy.year))

    # Filtro de seguridad por usuario
    es_admin = request.user.rol == 'ADMIN'
    filtro_usuario = {} if es_admin else {'usuario': request.user}

    meses_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }

    # Estructura para el desglose mensual
    datos_mensuales = []
    
    # Cálculos por cada mes en el rango
    for mes_num in range(mes_inicio, mes_fin + 1):
        def count_mes(model, campo_fecha):
            return model.objects.filter(
                activo=True,
                **{f"{campo_fecha}__month": mes_num, f"{campo_fecha}__year": anio},
                **filtro_usuario
            ).count()

        m_emergencias = count_mes(MorbilidadEmergencia, 'fecha_diagnostico')
        m_especialistas = count_mes(MorbilidadEspecialista, 'created_at')
        m_ecos = count_mes(MorbilidadEcosonograma, 'fecha')
        m_no_asistidos = count_mes(PacienteNoAsistido, 'fecha_cita')

        datos_mensuales.append({
            'mes_num': mes_num,
            'nombre': meses_nombres[mes_num],
            'emergencias': m_emergencias,
            'especialistas': m_especialistas,
            'ecos': m_ecos,
            'no_asistidos': m_no_asistidos,
            'total': m_emergencias + m_especialistas + m_ecos + m_no_asistidos
        })

    # Totales Consolidados (Suma de los meses)
    totales = {
        'emergencias': sum(d['emergencias'] for d in datos_mensuales),
        'especialistas': sum(d['especialistas'] for d in datos_mensuales),
        'ecosonogramas': sum(d['ecos'] for d in datos_mensuales),
        'no_asistidos': sum(d['no_asistidos'] for d in datos_mensuales),
    }

    # Desgloses Globales del Periodo (Especialidades y Ecos)
    def filtrar_periodo(model, campo_fecha):
        return model.objects.filter(
            activo=True,
            **{f"{campo_fecha}__month__gte": mes_inicio, f"{campo_fecha}__month__lte": mes_fin, f"{campo_fecha}__year": anio},
            **filtro_usuario
        )

    # --- Mejorar el Top Especialidades (Combinando Especialistas + Emergencias) ---
    # Usamos los mismos campos de fecha que en la tabla general para consistencia
    esp_data = MorbilidadEspecialista.objects.filter(
        activo=True,
        created_at__month__gte=mes_inicio, created_at__month__lte=mes_fin, created_at__year=anio,
        **filtro_usuario
    ).values('especialidad').annotate(total=Count('id'))

    emerg_count = MorbilidadEmergencia.objects.filter(
        activo=True,
        fecha_diagnostico__month__gte=mes_inicio, fecha_diagnostico__month__lte=mes_fin, fecha_diagnostico__year=anio,
        **filtro_usuario
    ).count()
    
    # Convertir a lista manejable
    desglose_esp = []
    for item in esp_data:
        desglose_esp.append({'especialidad': item['especialidad'], 'total': item['total']})
    
    if emerg_count > 0:
        desglose_esp.append({'especialidad': 'Emergencias', 'total': emerg_count})
    
    # Ordenar por el que tiene más pacientes
    desglose_esp = sorted(desglose_esp, key=lambda x: x['total'], reverse=True)

    # Corregir también el desglose de Ecos para que use el campo 'fecha' correctamente
    desglose_eco = MorbilidadEcosonograma.objects.filter(
        activo=True,
        fecha__month__gte=mes_inicio, fecha__month__lte=mes_fin, fecha__year=anio,
        **filtro_usuario
    ).values('tipo_eco').annotate(total=Count('id')).order_by('-total')

    # --- Cálculo de Insights para el Administrador ---
    mes_max = None
    max_valor = -1
    servicio_predominante = "N/A"
    
    if datos_mensuales:
        # Encontrar mes con más actividad
        mes_max_obj = max(datos_mensuales, key=lambda x: x['total'])
        mes_max = mes_max_obj['nombre']
        max_valor = mes_max_obj['total']
        
        # Encontrar servicio predominante en el total del periodo
        servicios_totales = {
            'Emergencias': totales['emergencias'],
            'Especialistas': totales['especialistas'],
            'Ecosonogramas': totales['ecosonogramas'],
            'No Asistidos': totales['no_asistidos']
        }
        servicio_predominante = max(servicios_totales, key=servicios_totales.get)

    # Calcular variación respecto al mes anterior (Tendencia)
    for i in range(len(datos_mensuales)):
        if i > 0:
            prev_total = datos_mensuales[i-1]['total']
            curr_total = datos_mensuales[i]['total']
            if prev_total == 0:
                variacion = 100 if curr_total > 0 else 0
            else:
                variacion = ((curr_total - prev_total) / prev_total) * 100
            datos_mensuales[i]['variacion'] = round(variacion, 1)
        else:
            datos_mensuales[i]['variacion'] = 0

    context = {
        'totales': totales,
        'datos_mensuales': datos_mensuales,
        'desglose_esp': desglose_esp,
        'desglose_eco': desglose_eco,
        'mes_inicio': mes_inicio,
        'mes_fin': mes_fin,
        'anio': anio,
        'meses_lista': sorted(meses_nombres.items()),
        'total_periodo': sum(totales.values()),
        'insight_mes_max': mes_max,
        'insight_max_valor': max_valor,
        'insight_servicio_top': servicio_predominante,
        'promedio_mensual': round(sum(totales.values()) / len(datos_mensuales), 1) if datos_mensuales else 0
    }

    return render(request, 'reportes/reporte_periodo.html', context)


@login_required
def exportar_reporte_periodo_excel_view(request):
    """Genera un archivo Excel multihonja usando openpyxl de forma nativa"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from io import BytesIO

    mes_inicio = int(request.GET.get('mes_inicio', 1))
    mes_fin = int(request.GET.get('mes_fin', 12))
    anio = int(request.GET.get('anio', timezone.now().year))
    
    filtro_usuario = {} if request.user.rol == 'ADMIN' else {'usuario': request.user}

    # Crear Libro
    wb = openpyxl.Workbook()
    
    # Estilos Profesionales
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def crear_hoja(nombre, model, campo_fecha, headers, campos):
        ws = wb.create_sheet(title=nombre)
        # Encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # Datos
        registros = model.objects.filter(
            activo=True,
            **{f"{campo_fecha}__month__gte": mes_inicio, 
               f"{campo_fecha}__month__lte": mes_fin, 
               f"{campo_fecha}__year": anio},
            **filtro_usuario
        ).order_by(campo_fecha)

        for r_idx, reg in enumerate(registros, 2):
            for c_idx, campo in enumerate(campos, 1):
                val = getattr(reg, campo, "")
                if callable(val): val = val()
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val))
                cell.border = thin_border

    # Hoja 1: Emergencias
    crear_hoja(
        "Emergencias", MorbilidadEmergencia, "fecha_diagnostico",
        ["Cédula", "Paciente", "Edad", "Sexo", "Médico", "Fecha"],
        ["cedula", "nombre_apellido", "edad", "sexo", "medico", "fecha_diagnostico"]
    )

    # Hoja 2: Especialistas
    crear_hoja(
        "Especialistas", MorbilidadEspecialista, "created_at",
        ["Paciente", "Edad", "Especialista", "Especialidad", "Diagnóstico"],
        ["nombre_apellido", "edad", "especialista", "especialidad", "diagnostico"]
    )

    # Hoja 3: Ecosonogramas
    crear_hoja(
        "Ecosonogramas", MorbilidadEcosonograma, "fecha",
        ["Paciente", "Cédula", "Tipo Eco", "Médico", "Diagnóstico", "Fecha"],
        ["nombre_apellido", "cedula", "tipo_eco", "medico", "diagnostico", "fecha"]
    )

    # Hoja 4: No Asistidos
    crear_hoja(
        "No Asistidos", PacienteNoAsistido, "fecha_cita",
        ["Paciente", "Edad", "Médico", "Especialidad", "Fecha Cita"],
        ["nombre_completo", "edad", "medico", "especialidad", "fecha_cita"]
    )

    # Eliminar la hoja por defecto creada por openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Preparar Respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_SICEME_{mes_inicio}_a_{mes_fin}_{anio}.xlsx"'
    return response
